import pandas as pd
import numpy as np
from pathlib import Path
from scipy import stats

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 1. PATH SETTINGS
# ============================================================
ROOT = Path(__file__).resolve().parents[1]

TRAIN_PATH = ROOT / "data" / "interim" / "train_raw.csv"
TEST_PATH = ROOT / "data" / "interim" / "internal_test_raw.csv"
EXTERNAL_PATH = ROOT / "data" / "interim" / "external_raw_aligned.csv"

OUT_DIR = ROOT / "output" / "distribution_check"
OUT_DIR.mkdir(parents=True, exist_ok=True)

LABEL_COL = "label"

# 明确删除的非预测性变量
# 注意：不要删除 time_of_TACE
DROP_EXACT_COLS = {
    "number",
    "source",
    "old_split",
}

# 只删除明显标识/日期变量，不放 "time"，避免误删 time_of_TACE
DROP_KEYWORDS = [
    "name",
    "姓名",
    "编号",
    "登记号",
    "住院号",
    "patient_id",
    "_id",
    "id_",
    "date",
    "初诊时间",
]

LABEL_MAP = {
    "cr": 1, "CR": 1,
    "pr": 1, "PR": 1,
    "sd": 0, "SD": 0,
    "pd": 0, "PD": 0,
    "response": 1,
    "nonresponse": 0,
    "non-response": 0,
}

# 强制指定连续变量
FORCE_CONTINUOUS = {
    "hemoglobin",
    "pt",
    "aptt",
    "tt",
    "alt",
    "ast",
    "alb",
    "tbil",
    "dbil",
    "ggt",
    "alp",
    "pa",
    "afp",
    "time_of_tace",
    "number_of_tumor",
    "diameter_of_tumor",
    "hbsag",
}

# 强制指定分类变量
FORCE_CATEGORICAL = {
    "pvtt",
    "combined_with_other_treatment",
    "tumor_location",
    "bclc_stage",
}

# 表1变量顺序，不加大类标题
FEATURE_ORDER = [
    "hemoglobin",
    "PT",
    "APTT",
    "TT",
    "ALT",
    "AST",
    "ALB",
    "TBIL",
    "DBIL",
    "GGT",
    "ALP",
    "PA",
    "AFP",
    "time_of_TACE",
    "PVTT",
    "number_of_tumor",
    "diameter_of_tumor",
    "HBsAg",
    "combined_with_other_treatment",
    "tumor_location",
    "BCLC_stage",
]


# ============================================================
# 2. HELPER FUNCTIONS
# ============================================================
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    统一列名：
    time of TACE -> time_of_TACE
    BCLC stage   -> BCLC_stage
    """
    df = df.copy()
    df.columns = [str(c).strip().replace(" ", "_") for c in df.columns]
    return df


def feature_key(col: str) -> str:
    return str(col).strip().replace(" ", "_").lower()


def display_name(col: str) -> str:
    """
    表格里显示用的变量名，可按论文习惯继续修改。
    """
    name_map = {
        "hemoglobin": "Hemoglobin",
        "PT": "PT",
        "APTT": "APTT",
        "TT": "TT",
        "ALT": "ALT",
        "AST": "AST",
        "ALB": "ALB",
        "TBIL": "TBIL",
        "DBIL": "DBIL",
        "GGT": "GGT",
        "ALP": "ALP",
        "PA": "PA",
        "AFP": "AFP",
        "time_of_TACE": "Time of TACE",
        "PVTT": "PVTT",
        "number_of_tumor": "Number of tumors",
        "diameter_of_tumor": "Diameter of tumor",
        "HBsAg": "HBsAg",
        "combined_with_other_treatment": "Combined with other treatment",
        "tumor_location": "Tumor location",
        "BCLC_stage": "BCLC stage",
    }
    return name_map.get(col, col.replace("_", " "))


def normalize_label(df: pd.DataFrame, name: str) -> pd.DataFrame:
    df = df.copy()

    if LABEL_COL not in df.columns:
        print(f"⚠️  {name}: label column not found.")
        print(list(df.columns))
        return df

    if df[LABEL_COL].dtype == "object":
        df[LABEL_COL] = df[LABEL_COL].astype(str).str.strip()
        df[LABEL_COL] = df[LABEL_COL].map(lambda x: LABEL_MAP.get(x, x))

    df[LABEL_COL] = pd.to_numeric(df[LABEL_COL], errors="coerce")

    if df[LABEL_COL].isna().any():
        raise ValueError(f"{name}: label contains unmapped or missing values.")

    df[LABEL_COL] = df[LABEL_COL].astype(int)
    return df


def drop_non_predictive(df: pd.DataFrame) -> pd.DataFrame:
    """
    删除 number/source/old_split 等非预测性变量。
    保留 time_of_TACE。
    """
    df = df.copy()
    drop_cols = []

    for col in df.columns:
        key = feature_key(col)

        if key in DROP_EXACT_COLS:
            drop_cols.append(col)
            continue

        for kw in DROP_KEYWORDS:
            if kw.lower() in key:
                drop_cols.append(col)
                break

    drop_cols = sorted(set(drop_cols))

    if drop_cols:
        print(f"Dropped non-predictive columns: {drop_cols}")
        df = df.drop(columns=drop_cols)

    return df


def label_distribution(df: pd.DataFrame, dataset_name: str) -> pd.DataFrame:
    vc = df[LABEL_COL].value_counts(dropna=False).sort_index()
    vp = df[LABEL_COL].value_counts(normalize=True, dropna=False).sort_index()

    return pd.DataFrame({
        "dataset": dataset_name,
        "label": vc.index,
        "count": vc.values,
        "proportion": vp.values,
    })


def infer_variable_type(df1: pd.DataFrame, df2: pd.DataFrame, col: str) -> str:
    key = feature_key(col)

    if key in FORCE_CONTINUOUS:
        return "continuous"

    if key in FORCE_CATEGORICAL:
        return "categorical"

    x_num = pd.to_numeric(df1[col], errors="coerce")
    y_num = pd.to_numeric(df2[col], errors="coerce")

    numeric_ratio = (x_num.notna().mean() + y_num.notna().mean()) / 2
    n_unique = pd.concat([x_num, y_num]).dropna().nunique()

    if numeric_ratio > 0.8 and n_unique > 5:
        return "continuous"

    return "categorical"


def is_approximately_normal(x: pd.Series) -> bool:
    """
    用 Shapiro-Wilk 初步判断正态性。
    样本量太小时不认为正态。
    """
    x = pd.to_numeric(x, errors="coerce").dropna()
    if len(x) < 3:
        return False

    try:
        return stats.shapiro(x).pvalue > 0.05
    except Exception:
        return False


def format_p(p: float) -> str:
    if pd.isna(p):
        return ""
    if p < 0.001:
        return "<0.001"
    return f"{p:.3f}"


def format_mean_sd(x: pd.Series) -> str:
    x = pd.to_numeric(x, errors="coerce").dropna()
    if len(x) == 0:
        return ""
    return f"{x.mean():.2f} ± {x.std(ddof=1):.2f}"


def format_median_iqr(x: pd.Series) -> str:
    x = pd.to_numeric(x, errors="coerce").dropna()
    if len(x) == 0:
        return ""
    return f"{x.median():.2f} ({x.quantile(0.25):.2f}, {x.quantile(0.75):.2f})"


def format_continuous_value(x: pd.Series, use_mean_sd: bool) -> str:
    if use_mean_sd:
        return format_mean_sd(x)
    return format_median_iqr(x)


def continuous_summary(group1: pd.DataFrame, group2: pd.DataFrame, feature: str):
    x = pd.to_numeric(group1[feature], errors="coerce").dropna()
    y = pd.to_numeric(group2[feature], errors="coerce").dropna()

    if len(x) == 0 or len(y) == 0:
        return None

    mean1, std1 = x.mean(), x.std(ddof=1)
    mean2, std2 = y.mean(), y.std(ddof=1)

    pooled_std = np.sqrt((std1**2 + std2**2) / 2)
    smd = (mean1 - mean2) / pooled_std if pooled_std > 0 else np.nan

    normal1 = is_approximately_normal(x)
    normal2 = is_approximately_normal(y)

    if normal1 and normal2:
        try:
            p_value = stats.ttest_ind(x, y, equal_var=False, nan_policy="omit").pvalue
        except Exception:
            p_value = np.nan
        test_name = "Welch t test"
        use_mean_sd = True
    else:
        try:
            p_value = stats.mannwhitneyu(x, y, alternative="two-sided").pvalue
        except Exception:
            p_value = np.nan
        test_name = "Mann–Whitney U"
        use_mean_sd = False

    return {
        "feature": feature,
        "type": "continuous",
        "group1_n": len(x),
        "group2_n": len(y),
        "group1_mean": mean1,
        "group2_mean": mean2,
        "group1_std": std1,
        "group2_std": std2,
        "group1_median": x.median(),
        "group2_median": y.median(),
        "group1_q1": x.quantile(0.25),
        "group2_q1": y.quantile(0.25),
        "group1_q3": x.quantile(0.75),
        "group2_q3": y.quantile(0.75),
        "SMD": abs(smd),
        "p_value": p_value,
        "test": test_name,
        "use_mean_sd": use_mean_sd,
    }


def smart_sort_levels(levels):
    """
    分类变量水平排序：
    数字优先按数值排，其余按字符串排。
    """
    def sort_key(x):
        s = str(x)
        try:
            return (0, float(s))
        except Exception:
            return (1, s)
    return sorted(levels, key=sort_key)


def categorical_summary(group1: pd.DataFrame, group2: pd.DataFrame, feature: str):
    x = group1[feature].fillna("Missing").astype(str)
    y = group2[feature].fillna("Missing").astype(str)

    tab = pd.crosstab(
        pd.Series(["group1"] * len(x) + ["group2"] * len(y), name="group"),
        pd.Series(list(x) + list(y), name=feature),
    )

    if tab.shape[1] <= 1:
        return None

    try:
        if tab.shape == (2, 2):
            chi2, chi_p, dof, expected = stats.chi2_contingency(tab.values)
            if (expected < 5).any() or (tab.values < 5).any():
                p_value = stats.fisher_exact(tab.values)[1]
                test_name = "Fisher exact"
            else:
                p_value = chi_p
                test_name = "Chi-square"
        else:
            p_value = stats.chi2_contingency(tab.values)[1]
            test_name = "Chi-square"
    except Exception:
        p_value = np.nan
        test_name = "NA"

    prop1 = x.value_counts(normalize=True)
    prop2 = y.value_counts(normalize=True)
    cats = smart_sort_levels(set(prop1.index) | set(prop2.index))
    max_abs_diff = max(abs(prop1.get(c, 0) - prop2.get(c, 0)) for c in cats)

    return {
        "feature": feature,
        "type": "categorical",
        "group1_n": len(x),
        "group2_n": len(y),
        "max_abs_prop_diff": max_abs_diff,
        "p_value": p_value,
        "test": test_name,
    }


def format_n_pct(n: int, total: int) -> str:
    if total <= 0:
        return ""
    return f"{n} ({100 * n / total:.1f}%)"


def categorical_level_rows(all_df: pd.DataFrame, group1: pd.DataFrame, group2: pd.DataFrame, feature: str):
    """
    生成分类变量各水平 n (%) 明细。
    """
    all_s = all_df[feature].fillna("Missing").astype(str)
    x = group1[feature].fillna("Missing").astype(str)
    y = group2[feature].fillna("Missing").astype(str)

    levels = smart_sort_levels(set(all_s.unique()) | set(x.unique()) | set(y.unique()))

    all_counts = all_s.value_counts(dropna=False)
    x_counts = x.value_counts(dropna=False)
    y_counts = y.value_counts(dropna=False)

    rows = []
    for lvl in levels:
        all_n = int(all_counts.get(lvl, 0))
        x_n = int(x_counts.get(lvl, 0))
        y_n = int(y_counts.get(lvl, 0))

        rows.append({
            "feature": feature,
            "level": lvl,
            "All": format_n_pct(all_n, len(all_s)),
            "Train": format_n_pct(x_n, len(x)),
            "Internal_test": format_n_pct(y_n, len(y)),
        })

    return rows


def compare_groups(df1: pd.DataFrame, df2: pd.DataFrame, name1: str, name2: str, out_prefix: str):
    common_cols = sorted(set(df1.columns) & set(df2.columns))
    common_cols = [c for c in common_cols if c != LABEL_COL]

    rows = []
    cat_rows = []

    for col in common_cols:
        if df1[col].isna().all() and df2[col].isna().all():
            continue

        var_type = infer_variable_type(df1, df2, col)

        if var_type == "continuous":
            res = continuous_summary(df1, df2, col)
            if res:
                res["group1"] = name1
                res["group2"] = name2
                rows.append(res)
        else:
            res = categorical_summary(df1, df2, col)
            if res:
                res["group1"] = name1
                res["group2"] = name2
                cat_rows.append(res)

    cont_df = pd.DataFrame(rows)
    cat_df = pd.DataFrame(cat_rows)

    if len(cont_df) > 0:
        cont_df = cont_df.sort_values("SMD", ascending=False)
        cont_df.to_csv(
            OUT_DIR / f"{out_prefix}_continuous.csv",
            index=False,
            encoding="utf-8-sig",
        )

    if len(cat_df) > 0:
        cat_df = cat_df.sort_values("max_abs_prop_diff", ascending=False)
        cat_df.to_csv(
            OUT_DIR / f"{out_prefix}_categorical.csv",
            index=False,
            encoding="utf-8-sig",
        )

    return cont_df, cat_df


def build_table1(all_df: pd.DataFrame, train_df: pd.DataFrame, test_df: pd.DataFrame,
                 cont_df: pd.DataFrame, cat_df: pd.DataFrame) -> pd.DataFrame:
    """
    生成论文风格 Table 1：
    Characteristics | All | Training set | Internal test set | P value
    不加大类标题。
    """
    rows = []

    cont_map = {}
    if len(cont_df) > 0:
        cont_map = {r["feature"]: r for _, r in cont_df.iterrows()}

    cat_map = {}
    if len(cat_df) > 0:
        cat_map = {r["feature"]: r for _, r in cat_df.iterrows()}

    available_features = [c for c in FEATURE_ORDER if c in all_df.columns]

    # 如果有未列在 FEATURE_ORDER 但仍保留的预测变量，也追加进去
    remaining = [
        c for c in all_df.columns
        if c not in available_features and c != LABEL_COL
    ]
    available_features += remaining

    for feature in available_features:
        var_type = infer_variable_type(train_df, test_df, feature)

        if var_type == "continuous":
            if feature not in cont_map:
                continue

            info = cont_map[feature]
            use_mean_sd = bool(info.get("use_mean_sd", False))

            rows.append({
                "Characteristics": display_name(feature),
                f"All\nN={len(all_df)}": format_continuous_value(all_df[feature], use_mean_sd),
                f"Training set\nN={len(train_df)}": format_continuous_value(train_df[feature], use_mean_sd),
                f"Internal test set\nN={len(test_df)}": format_continuous_value(test_df[feature], use_mean_sd),
                "P value": format_p(info["p_value"]),
            })

        else:
            if feature not in cat_map:
                continue

            info = cat_map[feature]

            # 分类变量主行，只放变量名和 P 值
            rows.append({
                "Characteristics": display_name(feature),
                f"All\nN={len(all_df)}": "",
                f"Training set\nN={len(train_df)}": "",
                f"Internal test set\nN={len(test_df)}": "",
                "P value": format_p(info["p_value"]),
            })

            # 分类变量水平行，缩进
            for lvl_row in categorical_level_rows(all_df, train_df, test_df, feature):
                rows.append({
                    "Characteristics": f"    {lvl_row['level']}",
                    f"All\nN={len(all_df)}": lvl_row["All"],
                    f"Training set\nN={len(train_df)}": lvl_row["Train"],
                    f"Internal test set\nN={len(test_df)}": lvl_row["Internal_test"],
                    "P value": "",
                })

    table1 = pd.DataFrame(rows)
    return table1


def style_table1_xlsx(xlsx_path: Path):
    wb = None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
    except Exception as e:
        print(f"Could not style Excel file: {e}")
        return

    ws = wb["Table 1"]

    thin = Side(style="thin", color="000000")
    border_bottom = Border(bottom=thin)

    # 页面基本设置
    ws.freeze_panes = "A3"

    # 标题行
    ws["A1"] = "Table 1. Baseline clinical characteristics of patients in the training and internal test sets"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # 表头在第2行
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_bottom

    # 内容格式
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(size=11)

    # 变量名列左对齐，其余居中
    for row in ws.iter_rows(min_row=3):
        row[0].alignment = Alignment(horizontal="left", vertical="center")
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置列宽
    widths = {
        "A": 32,
        "B": 24,
        "C": 24,
        "D": 24,
        "E": 12,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 最后一行加下边框
    max_row = ws.max_row
    for cell in ws[max_row]:
        cell.border = border_bottom

    wb.save(xlsx_path)


# ============================================================
# 3. MAIN
# ============================================================
def main():
    print("=" * 70)
    print("Distribution check and Table 1 generation")
    print("=" * 70)

    if not TRAIN_PATH.exists():
        raise FileNotFoundError(f"Cannot find {TRAIN_PATH}. Run scripts/split_dataset.py first.")
    if not TEST_PATH.exists():
        raise FileNotFoundError(f"Cannot find {TEST_PATH}. Run scripts/split_dataset.py first.")
    if not EXTERNAL_PATH.exists():
        raise FileNotFoundError(f"Cannot find {EXTERNAL_PATH}.")

    train = normalize_columns(pd.read_csv(TRAIN_PATH))
    internal_test = normalize_columns(pd.read_csv(TEST_PATH))
    external = normalize_columns(pd.read_csv(EXTERNAL_PATH))

    train = normalize_label(train, "train")
    internal_test = normalize_label(internal_test, "internal_test")
    external = normalize_label(external, "external")

    train = drop_non_predictive(train)
    internal_test = drop_non_predictive(internal_test)
    external = drop_non_predictive(external)

    # all 仅用于 Table 1 的 All 列，由 train + internal_test 合并得到
    internal_all = pd.concat([train, internal_test], axis=0, ignore_index=True)

    print("\nShapes after dropping non-predictive columns:")
    print("Internal all:", internal_all.shape)
    print("Train:", train.shape)
    print("Internal test:", internal_test.shape)
    print("External:", external.shape)

    print("\nColumns kept in train:")
    print(list(train.columns))

    if "number" in train.columns:
        print("⚠️ WARNING: number still exists in train columns.")
    else:
        print("✓ number removed.")

    if "time_of_TACE" in train.columns:
        print("✓ time_of_TACE kept.")
    else:
        print("⚠️ WARNING: time_of_TACE not found. Please check original column name.")

    # 保存 split summary
    split_summary = pd.DataFrame({
        "dataset": ["internal_all", "train_raw", "internal_test_raw"],
        "n": [len(internal_all), len(train), len(internal_test)],
    })
    split_summary.to_csv(
        OUT_DIR / "split_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # 保存 label 分布
    label_df = pd.concat(
        [
            label_distribution(train, "train"),
            label_distribution(internal_test, "internal_test"),
            label_distribution(external, "external"),
        ],
        ignore_index=True,
    )

    print("\nLabel distribution:")
    print(label_df)

    label_df.to_csv(
        OUT_DIR / "label_distribution.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # 列匹配情况
    col_report = pd.DataFrame({
        "column": sorted(set(train.columns) | set(internal_test.columns) | set(external.columns))
    })
    col_report["in_train"] = col_report["column"].isin(train.columns)
    col_report["in_internal_test"] = col_report["column"].isin(internal_test.columns)
    col_report["in_external"] = col_report["column"].isin(external.columns)
    col_report.to_csv(
        OUT_DIR / "column_overlap_report.csv",
        index=False,
        encoding="utf-8-sig",
    )

    print("\nComparing train vs internal_test...")
    cont1, cat1 = compare_groups(
        train,
        internal_test,
        "train",
        "internal_test",
        "train_vs_internal_test",
    )

    print("\nComparing train vs external...")
    cont2, cat2 = compare_groups(
        train,
        external,
        "train",
        "external",
        "train_vs_external",
    )

    # 生成分类变量各水平明细
    cat_level_rows = []
    for feature in cat1["feature"].tolist() if len(cat1) > 0 else []:
        cat_level_rows.extend(
            categorical_level_rows(internal_all, train, internal_test, feature)
        )

    cat_level_df = pd.DataFrame(cat_level_rows)
    cat_level_df.to_csv(
        OUT_DIR / "train_vs_internal_test_categorical_levels.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # 生成论文格式 Table 1
    table1 = build_table1(internal_all, train, internal_test, cont1, cat1)

    table1_csv = OUT_DIR / "table1_train_vs_internal_test_paper.csv"
    table1_xlsx = OUT_DIR / "table1_train_vs_internal_test_paper.xlsx"

    table1.to_csv(table1_csv, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(table1_xlsx, engine="openpyxl") as writer:
        # startrow=1 预留第1行作为表题
        table1.to_excel(writer, sheet_name="Table 1", index=False, startrow=1)

        cont1.to_excel(writer, sheet_name="Continuous detail", index=False)
        cat1.to_excel(writer, sheet_name="Categorical summary", index=False)
        cat_level_df.to_excel(writer, sheet_name="Categorical levels", index=False)
        label_df.to_excel(writer, sheet_name="Label distribution", index=False)
        split_summary.to_excel(writer, sheet_name="Split summary", index=False)

    style_table1_xlsx(table1_xlsx)

    print("\nSaved files to:", OUT_DIR)

    print("\nGenerated main files:")
    print(f"  - {OUT_DIR / 'split_summary.csv'}")
    print(f"  - {OUT_DIR / 'label_distribution.csv'}")
    print(f"  - {OUT_DIR / 'column_overlap_report.csv'}")
    print(f"  - {OUT_DIR / 'train_vs_internal_test_continuous.csv'}")
    print(f"  - {OUT_DIR / 'train_vs_internal_test_categorical.csv'}")
    print(f"  - {OUT_DIR / 'train_vs_internal_test_categorical_levels.csv'}")
    print(f"  - {table1_csv}")
    print(f"  - {table1_xlsx}")

    if len(cont1) > 0:
        print("\nTop continuous differences: train vs internal_test")
        print(cont1[["feature", "SMD", "p_value", "test"]].head(20))

    if len(cat1) > 0:
        print("\nCategorical differences: train vs internal_test")
        print(cat1[["feature", "max_abs_prop_diff", "p_value", "test"]])

    print("\n✅ DONE")
    print("\nImportant:")
    print("1. number/source/old_split are removed.")
    print("2. time_of_TACE is kept.")
    print("3. label is used only for label distribution, not Table 1 variables.")
    print("4. Table 1 does not include SMD in the main sheet.")
    print("5. Detailed SMD and categorical-level outputs are saved in separate sheets/files.")


if __name__ == "__main__":
    main()